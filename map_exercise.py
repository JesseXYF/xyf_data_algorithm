import openpyxl
import pickle


# 读取Excel文件
def readExel():
    filename = r'D:\Pydataproject\xyf_data_algorithm\Raw_Data\feature_dict_BDAI_map.xlsx'
    inwb = openpyxl.load_workbook(filename)  # 读文件

    sheetnames = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    ws = inwb.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容

    # 获取sheet的最大行数和列数
    rows = ws.max_row
    print(rows)
    cols = ws.max_column
    for r in range(2, rows+1):  # 从Excel表的第二行开始。
        # for c in range(1, cols):
        #     print(ws.cell(r, c).value)
        var_idx = ws.cell(r, 1).value
        valueSet_item = ws.cell(r, 2).value.split('\'')[1]
        if "demo2" == var_idx or "demo3" == var_idx or "demo4" == var_idx or "vital4" == var_idx \
                or "vital5" == var_idx or "vital6" == var_idx:
            Data.append(var_idx + valueSet_item)
            # print(var_idx)
        else:
            # print(valueSet_item)
            Data.append(var_idx)
        # if r == 56:
        #     break


Data = []
readExel()
cd = 'D:/Pydataproject/xyf_data_algorithm/result/'
# cdd = '/home/xzhang_sta/xyf/result/'
file = open(cd + "feature_dict_BDAI_map" + '.pkl', 'wb')
pickle.dump(Data, file)
file.close()
# print(Data)
# print(Data.index('demo2Y'))
